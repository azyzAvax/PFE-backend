from fastapi import UploadFile
import json
import os
from utils.excel_utils import save_temp_file, remove_temp_file, get_excel_sheet_names, extract_single_sheet_unstructured
from langchain_openai import ChatOpenAI
from langchain_core.prompts import PromptTemplate
from openpyxl import load_workbook
import logging

logger = logging.getLogger("uvicorn.error")

def init_llm(api_key: str) -> ChatOpenAI:
    """Initialize the LLM."""
    return ChatOpenAI(
        temperature=0,
        model_name="gpt-4o-mini",
        openai_api_key=api_key,
    )

def create_ddl_from_excel(file_path: str, sheet_name: str = None) -> str:
    """Generate DDL from an Excel sheet."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    schema_name = ws["C4"].value.strip() if ws["C4"].value else "default_schema"
    table_name = ws["C6"].value.strip() if ws["C6"].value else "default_table"
    table_comment = ws["C5"].value.strip() if ws["C5"].value else "No table comment provided"
    retention_time = 30

    ddl = f"CREATE OR REPLACE TABLE {schema_name}.{table_name} (\n"
    for row in ws.iter_rows(min_row=14, max_col=12, values_only=True):
        column_name = row[2]
        data_type = row[3]
        size = f"({row[4]})" if row[4] else ""
        not_null = "NOT NULL" if row[5] and row[5].strip().lower() == "yes" else ""
        default_value = f"DEFAULT {row[7]}" if row[7] else ""
        comment = f"COMMENT '{row[8].strip()}'" if row[8] else ""
        if not column_name or not data_type:
            continue
        ddl += f"    {column_name} {data_type}{size} {not_null} {default_value} {comment},\n"
    ddl = ddl.rstrip(",\n") + "\n)"
    ddl += f"\nDATA_RETENTION_TIME_IN_DAYS = {retention_time}\nCOMMENT = '{table_comment}';"
    return ddl

async def generate_ddl(file: UploadFile, env_vars: dict = None):
    """Generate DDL from an Excel file."""
    temp_path = await save_temp_file(file, file.filename)
    try:
        ddl = create_ddl_from_excel(temp_path)
        return ddl
    finally:
        remove_temp_file(temp_path)

async def get_sheet_names(file: UploadFile, env_vars: dict = None):
    """Retrieve sheet names from an Excel file."""
    temp_path = await save_temp_file(file, file.filename)
    try:
        sheet_names = get_excel_sheet_names(temp_path)
        return sheet_names
    finally:
        remove_temp_file(temp_path)

async def generate_ddl_from_sheets(file: UploadFile, sheets: str, env_vars: dict = None):
    """Generate DDLs for multiple sheets."""
    temp_path = await save_temp_file(file, file.filename)
    try:
        selected_sheets = json.loads(sheets)
        logger.info(f"Parsed sheets: {selected_sheets}")
        ddls = []
        for sheet in selected_sheets:
            ddl = create_ddl_from_excel(temp_path, sheet_name=sheet)
            ddls.append(ddl)
        return ddls
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON format for sheets: {sheets}")
        raise ValueError(f"Invalid sheets format. Expected a JSON-encoded list, e.g., '[\"Sheet1\", \"Sheet2\"]'. Error: {str(e)}")
    except Exception as e:
        logger.error(f"Error in generate_ddl_from_sheets: {str(e)}")
        raise
    finally:
        remove_temp_file(temp_path)

async def generate_ddl_from_design(file: UploadFile, sheets: str, env_vars: dict):
    """Generate DDL statements from design sheets."""
    temp_path = await save_temp_file(file, file.filename)
    try:
        logger.info(f"Received sheets parameter: {sheets}")
        
        if not sheets or sheets.strip() == "":
            raise ValueError("Sheets parameter cannot be empty")
        
        try:
            sheet_names = json.loads(sheets)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON format for sheets: {sheets}")
            raise ValueError(f"Invalid sheets format. Expected a JSON-encoded list, e.g., '[\"Sheet1\", \"Sheet2\"]'. Error: {str(e)}")
        
        if not isinstance(sheet_names, list):
            raise ValueError(f"Sheets must be a JSON-encoded list, got: {type(sheet_names)}")
        if not sheet_names:
            raise ValueError("Sheets list cannot be empty")

        ddl_results = []
        llm = init_llm(env_vars.get("OPENAI_API_KEY", os.getenv("OPENAI_API_KEY")))
        ddl_template = """
You are an AI expert in generating Snowflake DDL statements.

The following is the extracted structured data from a design Excel sheet describing a table.
Your job is to:
1. Understand the schema name, table name, and table comment.
2. Extract all column definitions dynamically (column name, type, size, nullability, default, comment).
3. Format the Snowflake SQL DDL accordingly.

Instructions:
- Do not assume fixed cell positions (like C4 or A1).
- Extract fields based on semantic meaning, not position.
- Use `NOT NULL` only if clearly marked.
- Use `DEFAULT value` only if defined.
- Use `COMMENT` only if a comment is available.
- Apply `DATA_RETENTION_TIME_IN_DAYS = 30` by default.

Return only the DDL statement. Do not include explanations or markdown.

Extracted Sheet: {extracted_data}
"""
        ddl_prompt = PromptTemplate.from_template(ddl_template)
        
        for sheet_name in sheet_names:
            extracted_data = extract_single_sheet_unstructured(temp_path, sheet_name)
            logger.info(f"Extracted data for sheet {sheet_name}: {extracted_data}")
            ddl_result = llm.invoke(ddl_prompt.format(extracted_data=extracted_data))
            ddl_results.append(ddl_result.content)
        return ddl_results
    except Exception as e:
        logger.error(f"Error in generate_ddl_from_design: {str(e)}")
        raise
    finally:
        remove_temp_file(temp_path)