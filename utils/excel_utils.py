import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from langchain_community.document_loaders import UnstructuredExcelLoader
from fastapi import UploadFile
import logging

logger = logging.getLogger("uvicorn.error")

async def save_temp_file(file: UploadFile, filename: str) -> str:
    """Save uploaded file temporarily and return its path."""
    temp_path = f"temp_{filename}"
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    logger.info(f"Saved temporary file to: {temp_path}")
    return temp_path

def remove_temp_file(file_path: str):
    """Remove temporary file."""
    if os.path.exists(file_path):
        os.remove(file_path)
        logger.info(f"Removed temporary file: {file_path}")

def get_excel_sheet_names(file_path: str) -> list:
    """Retrieve sheet names from an Excel file."""
    wb = load_workbook(file_path, data_only=True)
    return wb.sheetnames

def extract_single_sheet_unstructured(file_path: str, sheet_name: str) -> str:
    """Extract content from a specific Excel sheet."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        temp_path = f"temp_{sheet_name}.xlsx"
        df.to_excel(temp_path, index=False)
        loader = UnstructuredExcelLoader(temp_path, mode="elements")
        docs = loader.load()
        os.remove(temp_path)
        return "\n".join([doc.page_content for doc in docs])
    except Exception as e:
        logger.error(f"Error extracting sheet {sheet_name} from {file_path}: {str(e)}")
        raise